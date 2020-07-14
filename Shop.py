from datetime import datetime, timedelta
from datetime import date
import Error
from functools import partial
from idle_time import IdleMonitor
import Info
import keyboard
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
import numpy as np
import os
import PIL
from prettytable import PrettyTable
import random
import RSYSS
import requests
import sys
import sqlite3
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import time
import Till
import tempfile
from tkinter import messagebox
import tkinter.font as tkFont
import webbrowser
import xlsxwriter


import urllib
import urllib.parse
import io




conn = sqlite3.connect('Shop_Database.db')
c = conn.cursor()

c.execute('''CREATE TABLE IF NOT EXISTS CRJ(ID REAL,Day INTEGER,Month INTEGER,Year INTEGER,Time INTEGER, Date INTEGER, Description TEXT, Amount RAEL, Bank RAEL, Item TEXT, QTY TEXT, Payment_Type TEXT, Cashier TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS Sales (
                        "ID"    INTEGER,
                        "Day"   INTEGER,
                        "Month" INTEGER,
                        "Year"  INTEGER,
                        "Time"  INTEGER,
                        "Date"  INTEGER,
                        "Item"  TEXT,
                        "Price" INTEGER,
                        "Qty"   INTEGER,
                        "Weight" REAL)''')
c.execute('''CREATE TABLE IF NOT EXISTS Product_List(Code INT, Price REAL, Item TEXT, Cost_Price REAL)''')
c.execute('''CREATE TABLE IF NOT EXISTS Customer_List(Code REAL, Name TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS MenuS(Menu_Name TEXT, Menu_Number RAEL)''')
c.execute('''CREATE TABLE IF NOT EXISTS Cashiers(ID REAL, Name TEXT, Password TEXT, Permision REAL)''')
c.execute('''CREATE TABLE IF NOT EXISTS Scale(Code INTEGER, Name TEXT, Price_per_kg REAL)''')
c.execute('''CREATE TABLE IF NOT EXISTS Settings(ID REAL, Name TEXT, Value REAL, OTHER TEXT)''')
c.execute('''CREATE TABLE IF NOT EXISTS Tables(Item_PLU REAL, Item_Name TEXT, Item_Price INT, QTY INT, FPrice INT, Table_No INT)''')
c.execute('''CREATE TABLE IF NOT EXISTS Starters(Item_PLU REAL, Item_Name TEXT, Item_Price REAL, Menu INT)''')


def Create_Tables():
    c.execute('''CREATE TABLE IF NOT EXISTS CRJ(ID REAL,Day INTEGER,Month INTEGER,Year INTEGER,Time INTEGER, Date INTEGER, Description TEXT, Amount RAEL, Bank RAEL, Item TEXT, QTY TEXT, Payment_Type TEXT, Cashier TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS "Sales" (
                            "ID"	INTEGER,
                            "Day"	INTEGER,
                            "Month"	INTEGER,
                            "Year"	INTEGER,
                            "Time"	INTEGER,
                            "Date"      INTEGER,
                            "Item"	TEXT,
                            "Price"	INTEGER,
                            "Qty"	INTEGER,
                            "Weight" REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Product_List(Code INT, Price REAL, Item TEXT, Cost_Price REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Customer_List(Code REAL, Name TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS MenuS(Menu_Name TEXT, Menu_Number RAEL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Cashiers(ID REAL, Name TEXT, Password TEXT, Permision REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Scale(Code INTEGER, Name TEXT, Price_per_kg REAL)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Settings(ID REAL, Name TEXT, Value REAL, OTHER TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Tables(Item_PLU REAL, Item_Name TEXT, Item_Price INT, QTY INT, FPrice INT, Table_No INT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Starters(Item_PLU REAL, Item_Name TEXT, Item_Price REAL, Menu INT)''')
    def add_account(Name="admin",Password="1234",Perm=1):
        ID = random.randint(1,1000000000)
        c.execute('''INSERT INTO Cashiers(ID, Name, Password, Permision) VALUES(?, ? ,? ,?)''',(ID, Name, Password, Perm))
        conn.commit()
    try:
        c.execute("SELECT * FROM Cashiers WHERE Name=admin")
        for row in c.fetchall():
            print(row)
    except:
        add_account()
        f = open("Key.txt", "w")
        f.close()



Version = Info.i_version

#not Built In
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment

x = PrettyTable()




button_1 = "Beers & Ciders"
button_2 = "wine/per Glass"
button_3 = "Soft Drinks"
button_4 = "Gin"
button_5 = "Spirits"
button_6 = "Whiskey"
button_7 = "Brandy"




##try:
##    wb=load_workbook(filepath)
##    Previous_Value = 1
##except:
##    wb=Workbook()
##    # save workbook
##    wb.save(filepath)
##    Previous_Value = 1


##a1 = (row[0])
##b1 = 14 - (len(a))
##print (a1 + (" " * b1) + "|")



GEN = RSYSS
GEN.Config1()



TimeDate = (time.strftime("%H:%M--%d/%m/%Y"))

DEF1 = "0000"

NAME1 = "Admin"
PASS1 = "99562281"
NAME2 = "Cashier 1"
PASS2 = DEF1
NAME3 = "Cashier 2"
PASS3 = DEF1
Time2 = (time.strftime("%H:%M"))
DATE = (time.strftime("%d-%m-%Y"))

print (time.strftime("%H:%M"))
print (time.strftime("%d/%m/%Y"))
print (random.randint(1,1000000000))



##Create_Tables()
##try:
##    c.execute("SELECT * FROM Settings")
##    for row in c.fetchall():
####        print("H")
##        pass
##except:
##    SET.FIRSTTIME()

##Bbb1 = "1"
##Bbb2 = "1"
##Bbb3 = "1"
##c.execute('''INSERT INTO CRJ(PLU, Name, Price) VALUES(?, ? ,?)''',(Bbb1, Bbb2, Bbb3))
##conn.commit()


global LEVL

LEVL = 1

global fontStyle
global fontStyle2
global fontStyle3

##try:
##    c.execute("SELECT * FROM Settings WHERE ID=1")
##    for row in c.fetchall():
##        fontStyle = tkFont.Font(size=int(row[2]))
##except:
##    fontStyle = tkFont.Font(size=int(12))
##
##try:
##    c.execute("SELECT * FROM Settings WHERE ID=2")
##    for row in c.fetchall():
##        fontStyle2 = tkFont.Font(size=int(row[2]))
##except:
##    fontStyle2 = tkFont.Font(size=int(12))
##
##try:
##    c.execute("SELECT * FROM Settings WHERE ID=3")
##    for row in c.fetchall():
##        fontStyle3 = tkFont.Font(size=int(row[2]))
##except:
##    fontStyle3 = tkFont.Font(size=int(12))


def FTW():
    c.execute('''DROP TABLE Settings''')
    conn.commit()
    SET.FIRSTTIME()

def LEV1(Logged_In):
    global LEVL
    LEVL = 1
    RUN1(Logged_In)

def LEV2(Logged_In):
    global LEVL
    LEVL = 2
    RUN1(Logged_In)

def LEV3(Logged_In):
    global LEVL
    LEVL = 3
    RUN1(Logged_In)



def RUN1(Logged_In):
    Page1 = Tk()
    Page1.title("Shop Database")
    Page1.configure(background="#BEBEBE")
    Page1.attributes("-fullscreen", True)
    Page1.attributes("-topmost", True)

##    try:
##        c.execute("SELECT * FROM Settings WHERE ID=1")
##        for row in c.fetchall():
##            fontStyle = tkFont.Font(root=Page1, size=int(row[2]))
##    except:
##        fontStyle = tkFont.Font(root=Page1, size=int(12))
##
##    try:
##        c.execute("SELECT * FROM Settings WHERE ID=2")
##        for row in c.fetchall():
##            fontStyle2 = tkFont.Font(root=Page1, size=int(row[2]))
##    except:
##        fontStyle2 = tkFont.Font(root=Page1, size=int(12))
##
##    try:
##        c.execute("SELECT * FROM Settings WHERE ID=3")
##        for row in c.fetchall():
##            fontStyle3 = tkFont.Font(root=Page1, size=int(row[2]))
##    except:
##        fontStyle3 = tkFont.Font(root=Page1, size=int(12))



#=========================================================#Add_User#==================================================================

    def Add_User():
        Add_user = Toplevel()
        Add_user.title("Shop Database")
        Add_user.configure(background="#E9E9E9")
        Add_user.geometry("800x350+253+125")
        Add_user.transient([Page1])

##        Add_user.attributes("-topmost", True)

        Label(Add_user, text="Name", bd=2).grid(row=0,column=0,pady=2,sticky='e')
        User_Name = Entry(Add_user, bd=2)
        User_Name.grid(row=0,column=1,pady=2)

        Label(Add_user, text="Password", bd=2).grid(row=1,column=0,pady=2,sticky='e')
        User_Password = Entry(Add_user, bd=2)
        User_Password.grid(row=1,column=1,pady=2)

        def sel():
            PERM = int(var.get())

        F9 = Frame(Add_user, width=50, bg="#E9E9E9", relief="raise")
        F9.grid_propagate(0)
        F9.grid(row=0,column=3)

        var = IntVar()
        R1 = Radiobutton(Add_user, text="Admin", variable=var, value=1,
                          command=sel)
        R1.grid(row=0,column=4,pady=2)

        R2 = Radiobutton(Add_user, text="Manager", variable=var, value=2,
                          command=sel)
        R2.grid(row=1,column=4,pady=2)

        R3 = Radiobutton(Add_user, text="Cashier", variable=var, value=3,
                          command=sel)
        R3.grid(row=2,column=4,pady=2)

        def Commit_user():
            if len(User_Name.get()) <= 0:
                MsgBox_004 = messagebox.showerror ('ERROR',Error.Error_004,icon = Error.Error_icon)
            elif len(User_Password.get()) <= 0:
                MsgBox_005 = messagebox.showerror ('ERROR',Error.Error_005,icon = Error.Error_icon)
            elif User_Role == None:
                MsgBox_006 = messagebox.showerror ('ERROR',Error.Error_006,icon = Error.Error_icon)
            else:
                Till.add_account(User_Name.get(),User_Password.get(),PERM)
                Add_user.destroy()
                Add_User()


        Button(Add_user, text="Add", width=20, height=1, fg="white", bg="green", command=Commit_user, bd=2).grid(row=2,column=1)

#=========================================================#Remove_User#==================================================================

    def Delete_User():
        Delete_user = Toplevel()
        Delete_user.title("Shop Database")
        Delete_user.configure(background="#E9E9E9")
        Delete_user.geometry("800x350+253+125")
        Delete_user.transient([Page1])

        LbName1 = Listbox(Delete_user)
        c.execute("SELECT * FROM Cashiers")
        for row in c.fetchall():
    ##        print(row)
    ##        print(row[1])
            Name = row[1]
            LbName1.insert(1, Name)

        def Delete_user_C():
            try:
                User = LbName1.get(LbName1.curselection())
            except:
                MsgBox_003 = messagebox.showerror ('ERROR',Error.Error_003,icon = Error.Error_icon)
            print(User)
            Till.delete_account(User)
            Delete_user.destroy()
            Delete_User()


        Button(Delete_user, text="Delete", width=20, height=1, fg="white", bg="green", command=Delete_user_C, bd=2).grid(row=2,column=0)
        LbName1.grid(row=0,column=0)

#=========================================================#Add_To_Scale#==================================================================

    def Add_Scale():
        Add_scale = Toplevel()
        Add_scale.title("Shop Database")
        Add_scale.configure(background="#E9E9E9")
        Add_scale.geometry("800x350+253+125")
        Add_scale.transient([Page1])

##        Add_user.attributes("-topmost", True)

        Label(Add_scale, text="Code", bd=2).grid(row=0,column=0,pady=2,sticky='e')
        Label(Add_scale, text="Example: the full barcode of the item you want to add '2 100017 796304' with no spaces", bd=2).grid(row=0,column=2,pady=2,sticky='e')
        Scale_Code = Entry(Add_scale, bd=2)
        Scale_Code.grid(row=0,column=1,pady=2)

        Label(Add_scale, text="Name", bd=2).grid(row=1,column=0,pady=2,sticky='e')
        Scale_Name = Entry(Add_scale, bd=2)
        Scale_Name.grid(row=1,column=1,pady=2)

        Label(Add_scale, text="Price Per KG", bd=2).grid(row=2,column=0,pady=2,sticky='e')
        Scale_PPK = Entry(Add_scale, bd=2)
        Scale_PPK.grid(row=2,column=1,pady=2)


        def Add_Scale_C():
            if len(str(Scale_Code.get())) <= 0 or len(Scale_Code.get()) < 13:
                MsgBox_009 = messagebox.showerror ('ERROR',Error.Error_009,icon = Error.Error_icon)
            elif len(Scale_Name.get()) <= 0:
                MsgBox_004 = messagebox.showerror ('ERROR',Error.Error_004,icon = Error.Error_icon)
            elif len(Scale_PPK.get()) <= 0:
                MsgBox_013 = messagebox.showerror ('ERROR',Error.Error_013,icon = Error.Error_icon)
            else:
                try:
                    Scale_Code_d = int(Scale_Code.get())
                except:
                    MsgBox_012 = messagebox.showerror ('ERROR',Error.Error_012,icon = Error.Error_icon)
                try:
                    Scale_PPK_d = float(Scale_PPK.get())
                except:
                    MsgBox_014 = messagebox.showerror ('ERROR',Error.Error_014,icon = Error.Error_icon)
                if len(Scale_Code.get()) == 13:
                    s1,s2,s3,s4,s5,s6,s7,s8,s9,s10,s11,s12,s13 = str(Scale_Code.get())
                    Code_1 = s1+s2+s3+s4+s5+s6
                    Price_1 = s8+s9+s10+ "." +s11+s12
                    Till.add_scale(Code_1,Scale_Name.get(),Scale_PPK_d)
                    Add_scale.destroy()
                    Add_Scale()
                else:
                    MsgBox_014 = messagebox.showerror ('ERROR',Error.Error_014,icon = Error.Error_icon)

        Button(Add_scale, text="Add", width=20, height=1, fg="white", bg="green", command=Add_Scale_C, bd=2).grid(row=3,column=1)


#=========================================================#Remove_from_scale#==================================================================

    def Delete_Scale():
        Delete_scale = Toplevel()
        Delete_scale.title("Shop Database")
        Delete_scale.configure(background="#E9E9E9")
        Delete_scale.geometry("800x350+253+125")
        Delete_scale.transient([Page1])

        LbName1 = Listbox(Delete_scale)
        c.execute("SELECT * FROM Scale")
        for row in c.fetchall():
    ##        print(row)
    ##        print(row[1])
            Name = row[1]
            LbName1.insert(1, Name)

        def Delete_user_C():
            try:
                Name_scale = LbName1.get(LbName1.curselection())
            except:
                MsgBox_003 = messagebox.showerror ('ERROR',Error.Error_003,icon = Error.Error_icon)
            print(Name_scale)
            Till.delete_scale(Name_scale)
            Delete_scale.destroy()
            Delete_Scale()


        Button(Delete_scale, text="Delete", width=20, height=1, fg="white", bg="green", command=Delete_user_C, bd=2).grid(row=2,column=0)
        LbName1.grid(row=0,column=0)


#=========================================================#Help#==================================================================

    def Help():
        new = 2
        webbrowser.open("https://www.node-s.co.za/products/easypos/help",new=new)
##        webbrowser.open("file://C:/Users/conno/OneDrive/Desktop/Shop/Python totaurials.docx",new=new)

    def Open_Yoco():
        new = 2
        webbrowser.open("http://referral.yoco.com/DtvrN",new=new)
##        webbrowser.open("file://C:/Users/conno/OneDrive/Desktop/Shop/Python totaurials.docx",new=new)

    def Donate_Yoco():
        new = 2
        webbrowser.open("https://pay.yoco.com/node-s?reference=Donate",new=new)
##        webbrowser.open("file://C:/Users/conno/OneDrive/Desktop/Shop/Python totaurials.docx",new=new)


#=========================================================#restorant items#===========================================================

    def Add_Menu1():
        PageAM = Toplevel()
        PageAM.title("Shop Database")
        PageAM.configure(background="#E9E9E9")
        PageAM.geometry("800x350+253+125")
        PageAM.transient([Page1])



        Label(PageAM, text="Name", bd=2).place(x=1,y=1)
        Label(PageAM, text="Price", bd=2).place(x=1,y=26)
        Label(PageAM, text="Delete PLU", bd=2).place(x=1,y=90)
        Label(PageAM, text="PLU", bd=2).place(x=280,y=1)

        AM1 = Entry(PageAM, bd=2)
        AM1.place(x=100,y=1)
##        AM1.insert(0,"NULL")
        AM2 = Entry(PageAM, bd=2)
        AM2.place(x=100,y=26)
##        AM2.insert(0,"NULL")
        AM3 = Entry(PageAM, bd=2)
        AM3.place(x=100,y=90)

        AM4 = Entry(PageAM, bd=2)
        AM4.place(x=310,y=1)

        var4 = IntVar()
        var4.set(1)

        def ER():
            print(var4.get())

        Radio_x = 500
        X1 = 250
        X2 = 1
        button_number = 1

        Current_value = 1
        print('Previous_Value: ' + str(Previous_Value))

        def Value_update_1(Name,Valuee):
            f3= open("Previous_Value.txt","w")
            Current_value = Valuee
            Previous_Value = Current_value
            f3.write(str(Current_value))
            print('Previous_Value_after_Button: ' + str(Previous_Value))
            print('Current_value: ' + str(Current_value))
            f3.close()
            Button(PageAM, text="Add", width=20, height=1, fg="white", bg="green", command=partial(Add_Menu2,Current_value), bd=2).place(x=1,y=55)




        c.execute("SELECT * FROM MenuS ORDER BY Menu_Number ASC")
        for row in c.fetchall():
            row_0 = (row[0])
            row_1 = (row[1])
            if button_number == 1:
                R1 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=1, command=partial(Value_update_1,row_0,row_1))
                R1.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 2:
                R2 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=2, command=partial(Value_update_1,row_0,row_1))
                R2.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 3:
                R3 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=3, command=partial(Value_update_1,row_0,row_1))
                R3.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 4:
                R4 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=4, command=partial(Value_update_1,row_0,row_1))
                R4.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 5:
                R5 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=5, command=partial(Value_update_1,row_0,row_1))
                R5.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 6:
                R6 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=6, command=partial(Value_update_1,row_0,row_1))
                R6.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 7:
                R7 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=7, command=partial(Value_update_1,row_0,row_1))
                R7.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30

            elif button_number == 8:
                R8 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=8, command=partial(Value_update_1,row_0,row_1))
                R8.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 9:
                R9 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=9, command=partial(Value_update_1,row_0,row_1))
                R9.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 10:
                R10 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=10, command=partial(Value_update_1,row_0,row_1))
                R10.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 11:
                R11 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=11, command=partial(Value_update_1,row_0,row_1))
                R11.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 12:
                R12 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=12, command=partial(Value_update_1,row_0,row_1))
                R12.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 13:
                R13 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=13, command=partial(Value_update_1,row_0,row_1))
                R13.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 14:
                R14 = Radiobutton(PageAM, text=(row[0]), variable=var4, value=14, command=partial(Value_update_1,row_0,row_1))
                R14.place(x=Radio_x,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30




##        R1 = Radiobutton(PageAM, text=button_1, variable=var4, value=1, command=ER)
##        R1.place(x=Radio_x,y=1)
##
##        R2 = Radiobutton(PageAM, text=button_2, variable=var4, value=2, command=ER)
##        R2.place(x=Radio_x,y=30)
##
##        R3 = Radiobutton(PageAM, text=button_3, variable=var4, value=3, command=ER)
##        R3.place(x=Radio_x,y=60)
##
##        R4 = Radiobutton(PageAM, text=button_4, variable=var4, value=4, command=ER)
##        R4.place(x=Radio_x,y=90)
##
##        R5 = Radiobutton(PageAM, text=button_5, variable=var4, value=5, command=ER)
##        R5.place(x=Radio_x,y=120)
##
##        R6 = Radiobutton(PageAM, text=button_6, variable=var4, value=6, command=ER)
##        R6.place(x=Radio_x,y=150)
##
##        R7 = Radiobutton(PageAM, text=button_7, variable=var4, value=7, command=ER)
##        R7.place(x=Radio_x,y=180)



        def Reset2():
            if len(AM3.get()) <= 0:
                MsgBox_009 = messagebox.showerror ('ERROR',Error.Error_009,icon = Error.Error_icon)
            else:
                MsgBox = messagebox.askquestion ('Warning','Are you sure you want save this Item',icon = 'warning')
                if MsgBox == 'yes':
                    B1 = AM3.get()
                    c.execute('''DELETE FROM Starters WHERE Item_PLU=?''',(B1,))
                    conn.commit()
                    PageAM.destroy()
        def Add_Menu2(Current_value):
            print("Add current: " + str(Current_value))
            Bbb1 = AM4.get()
            Bbb2 = (AM1.get()).replace (" ", "_")
            Bbb3 = str(AM2.get()) + " "
            Bbb4 = str(Current_value)
            print(Bbb4)
            if len(AM1.get()) <= 0:
                MsgBox_004 = messagebox.showerror ('ERROR',Error.Error_004,icon = Error.Error_icon)
            elif len(AM2.get()) <= 0:
                MsgBox_007 = messagebox.showerror ('ERROR',Error.Error_007,icon = Error.Error_icon)
            elif len(AM4.get()) <= 0:
                MsgBox_008 = messagebox.showerror ('ERROR',Error.Error_008,icon = Error.Error_icon)
            else:
                MsgBox = messagebox.askquestion ('Warning','Are you sure you want save this Item',icon = 'warning')
                if MsgBox == 'yes':
                    GEN.ItemAdd(Bbb1, Bbb2, Bbb3, Bbb4)
                    PageAM.destroy()

        def delete_all():
            MsgBox = messagebox.askquestion ('Warning','Are you sure you want to Delete All\nThis will delete everything and can not be undone',icon = 'warning')
            if MsgBox == 'yes':
                Lb1.delete(1, 200)
                c.execute('''DELETE FROM Starters''')
                conn.commit()
                PageAM.destroy()

        Label(PageAM, text="Add", bd=2, width=20, height=1, bg='yellow', fg='red').place(x=1,y=55)
##        Button(PageAM, text="Add", width=20, height=1, fg="white", bg="green", command=partial(Add_Menu2,Current_value), bd=2).place(x=1,y=55)
        Button(PageAM, text="delete", width=20, height=1, fg="white", bg="green", command=Reset2, bd=2).place(x=1,y=120)
        Button(PageAM, text="delete all", width=20, height=1, fg="Black", bg="Red", command=delete_all, bd=2).place(x=1,y=180)




#==========================================================#Income graph#=======================================================================

    def GRAPH():
        PCRJ2 = Toplevel()
        PCRJ2.title("Shop Database")
        PCRJ2.configure(background="green")
        PCRJ2.geometry("1480x1080")

#==========================================================#CRJ#=======================================================================

    def CRJ1():
##        GRAPH()
        def Reset1():
            MsgBox = messagebox.askquestion ('Warning','Are you sure you want clear the sales History',icon = 'warning')
            if MsgBox == 'yes':
                c.execute('''DROP TABLE CRJ''')
                conn.commit()
                c.execute('''DROP TABLE Sales''')
                conn.commit()
                Create_Tables()
                PCRJ.destroy()
                CRJ1()

        sqlite3.connect('Shop_Database.db')
        DATE2 = (time.strftime("%d/%m/%Y"))
        PCRJ = Toplevel()
        PCRJ.title("Shop Database")
        PCRJ.configure(background="#E9E9E9")
        PCRJ.geometry("1530x1080+0+0")
        PCRJ.transient([Page1])


        Button(PCRJ, text="Clear", width=12, fg="white", bg="red", command=Reset1, bd=2).place(x=1,y=1)

        text = Text(PCRJ, width=180)

        BZ = PrettyTable()
        c.execute("SELECT ID, Date,Time,Description,Amount,Bank,Item,Payment_Type,Cashier FROM CRJ")
        for row in c.fetchall():
            BZ.field_names = ["ID", "Date", "Time", "Description", "Amount paid", "Bank", "Item","Payment Type","Cashier"]
            BZB = (row)
            BZ.add_row((BZB))
        print(BZ)
        text.insert(INSERT, BZ)

        text.pack(side="top")

        text_I = Text(PCRJ, width=130)

        Query_Date_T = time.strftime("%d-%m-%Y")
        date1 = datetime.strptime(Query_Date_T, "%d-%m-%Y")
        modified_date = date1 + timedelta(days=-1)
        Query_Date_Y = datetime.strftime(modified_date, "%d-%m-%Y")

        BZ = PrettyTable()

        Total_Tenderd_Cash_Yesterday = 0.0
        Total_Meat_Yesterday = 0.0
        Total_Meat_Today = 0.0
        Total_Tenderd_Cash_Today = 0.0


        c.execute("SELECT * FROM Sales Where Date=?",(Query_Date_Y,))
        for row in c.fetchall():
            Total_Meat_Yesterday = float(Total_Meat_Yesterday) + float(row[9])
            Total_Tenderd_Cash_Yesterday = float(Total_Tenderd_Cash_Yesterday) + float(row[7])

        c.execute("SELECT * FROM Sales Where Date=?",(Query_Date_T,))
        for row in c.fetchall():
            Total_Meat_Today = float(Total_Meat_Today) + float(row[9])
            Total_Tenderd_Cash_Today = float(Total_Tenderd_Cash_Today) + float(row[7])

        BZ.field_names = ["Total Meat In KG [Yesterday]", "Total Cash Recived [Yesterday]", "Total Meat In KG [Today]", "Total Cash Recived [Today]",]
        BZ.add_row([Total_Meat_Yesterday, Total_Tenderd_Cash_Yesterday, Total_Meat_Today, Total_Tenderd_Cash_Today])
        text_I.insert(INSERT, BZ)

        text_I.pack(side="bottom")


#==========================================================#Restorant System#===================================================================

    def RSYS():
        GEN.RSYS(Logged_In)

#===========================================================#Delete product#===============================================================
    def Del_product():
        sqlite3.connect('Shop_Database.db')
        Page21 = Toplevel()
        Page21.title("Item Database")
        Page21.configure(background="#E9E9E9")
        Page21.geometry("800x350+253+125")
        Page21.transient([Page1])


        def Cancel():
            Page21.destroy()

        Label(Page21, text="Code", bd=2).place(x=1,y=1)
        e1 = Entry(Page21, bd=2)
        e1.place(x=100,y=1)

        Button(Page21, text="Cancel", width=12, fg="white", bg="red", command=Cancel, bd=2).place(x=90,y=30)

        def data_entry1():
            if len(e1.get()) <= 0:
                MsgBox_008 = messagebox.showerror ('ERROR',Error.Error_008,icon = Error.Error_icon)
            else:
                MsgBox = messagebox.askquestion ('Warning','Are you sure you want Delete this Item',icon = 'warning')
                if MsgBox == 'yes':
                    sqlite3.connect('Shop_Database.db')
                    B1 = e1.get()
                    c.execute('''DELETE FROM Product_List WHERE Code=?''',(B1,))
                    conn.commit()
                    Page21.destroy()
                    Del_product()


        Button(Page21, text="Delete", width=12, fg="white", bg="green", command=data_entry1, bd=2).place(x=1,y=30)



#==============================================================#product list#============================================================

    def P_List():
        Page22 = Toplevel()
        Page22.title("Item Database")
        Page22.configure(background="#E9E9E9")
        Page22.geometry("800x350+253+125")
        Page22.transient([Page1])

        text4 = Text(Page22)
        text4.insert(INSERT, "|Code        |Price |Name                 |\n")

        c.execute("SELECT * FROM Product_List")
        DA22 = c.fetchall()
        for row in DA22:
            text4.insert(INSERT, (row))
            text4.insert(INSERT, "\n")
        text4.place(x=1,y=1)



#================================================================#Add Product#==========================================================


    def New_product():
        sqlite3.connect('Shop_Database.db')
        Page2 = Toplevel()
        Page2.title("Item Database")
        Page2.configure(background="#E9E9E9")
        Page2.geometry("800x350+253+125")
        Page2.transient([Page1])

        def Cancel():
            Page2.destroy()

        Label(Page2, text="Code", bd=2).place(x=1,y=40)
        Label(Page2, text="Price", bd=2).place(x=1,y=20)
        Label(Page2, text="Item Name", bd=2).place(x=1,y=1)
        Label(Page2, text="Cost Price", bd=2).place(x=250,y=1)

        e1 = Entry(Page2, bd=2)
        e1.place(x=100,y=1)
        e2 = Entry(Page2, bd=2)
        e2.place(x=100,y=20)
        e3 = Entry(Page2, bd=2)
        e3.place(x=100,y=40)
        e33 = Entry(Page2, bd=2)
        e33.place(x=320,y=1)

        Button(Page2, text="Cancel", width=12, fg="white", bg="red", command=Cancel, bd=2).place(x=90,y=70)

        def data_entry2():
            sqlite3.connect('Shop_Database.db')
            B1 = e1.get()
            B2 = e2.get()
            B3 = e3.get()
            B4 = e33.get()
            if len(e1.get()) <= 0:
                MsgBox_009 = messagebox.showerror ('ERROR',Error.Error_009,icon = Error.Error_icon)
            elif len(e2.get()) <= 0:
                MsgBox_007 = messagebox.showerror ('ERROR',Error.Error_007,icon = Error.Error_icon)
            elif len(e3.get()) <= 0:
                MsgBox_010 = messagebox.showerror ('ERROR',Error.Error_010,icon = Error.Error_icon)
            elif len(e33.get()) <= 0:
                MsgBox_011 = messagebox.showerror ('ERROR',Error.Error_011,icon = Error.Error_icon)
            else:
                MsgBox = messagebox.askquestion ('Warning','Are you sure you want save this Item',icon = 'warning')
                if MsgBox == 'yes':
                    c.execute('''INSERT INTO Product_List(Code, Price, Item, Cost_Price) VALUES(?, ? ,? ,?)''',(B3, B2, B1, B4))
                    conn.commit()
                    Page2.destroy()
                    New_product()
        Button(Page2, text="Add", width=12, fg="white", bg="green", command=data_entry2, bd=2).place(x=1,y=70)



#=================================================#Shop cart#==================================================================================

    def Cart1():
        Till.Login_Page()


    def ESC():
        filepath = os.path.expanduser('~/Documents/EasyPOS_CRJ_' + str(time.strftime("%d_%m_%Y")) + '.xlsx')
        try:
            wb=load_workbook(filepath)
        except:
            wb=Workbook()

        print(filepath)
        try:
            sheet = wb[(time.strftime("%H-%M"))]
        except KeyError:
            sheet = wb.create_sheet((time.strftime("%H-%M")))
        Header = ('ID','Day','Month','Year','Time','Date','Description','Amount','Bank','Item','QTY','Payment_Type','Cashier')
        sheet.append(Header)
        sheet.column_dimensions['A'].width = float('11.67')
        sheet.column_dimensions['B'].width = float('6.56')
        sheet.column_dimensions['C'].width = float('6.56')
        sheet.column_dimensions['D'].width = float('5.44')
        sheet.column_dimensions['E'].width = float('5.67')
        sheet.column_dimensions['F'].width = float('11.22')
        sheet.column_dimensions['G'].width = float('9.67')
        sheet.column_dimensions['J'].width = 60
        sheet.column_dimensions['K'].width = float('4')


        c.execute("SELECT * FROM CRJ")
        for row in c.fetchall():
            sheet.append(row)

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

        wb.save(filepath)
        Page1.iconify()
        os.startfile(filepath)


#======================================================================================================================================================================
#========================================================================#Home Layout#=================================================================================
#======================================================================================================================================================================
    def donothing():
            pass
    def Page1_close():
        Page1.destroy()

    Page1.update_idletasks()
    if LEVL == 1:

        menubar = Menu(Page1)
        filemenu = Menu(menubar, tearoff=0)
        filemenu.add_command(label="Exit", command=Page1_close)
        menubar.add_cascade(label="File", menu=filemenu)

        Usermenu = Menu(menubar, tearoff=0)
        Usermenu.add_command(label="User List", command=donothing)
        Usermenu.add_separator()
        Usermenu.add_command(label="Add User", command=Add_User)
        Usermenu.add_command(label="Delete User", command=Delete_User)
        Usermenu.add_separator()
        Usermenu.add_command(label="Log(Login)", command=donothing)
        menubar.add_cascade(label="User's", menu=Usermenu)

        Posmenu = Menu(menubar, tearoff=0)
        Posmenu.add_command(label="Product List", command=P_List)
        Posmenu.add_separator()
        Posmenu.add_command(label="New Product", command=New_product)
        Posmenu.add_command(label="Delete Product", command=Del_product)
        Posmenu.add_separator()
        Posmenu.add_command(label="Add Scale Item", command=Add_Scale)
        Posmenu.add_command(label="Delete Scale Item", command=Delete_Scale)
        menubar.add_cascade(label="POS System", menu=Posmenu)

        Hotel_Mode = IntVar()
        Restaurantmenu = Menu(menubar, tearoff=0)
        Restaurantmenu.add_command(label="Item List", command=donothing)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_command(label="New Item", command=Add_Menu1)
        Restaurantmenu.add_command(label="Delete Item", command=Add_Menu1)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_command(label="Add Menu category", command=donothing)
        Restaurantmenu.add_command(label="Remove Menu category", command=donothing)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_checkbutton(label="Hotel Mode", onvalue=1, offvalue=0, variable=Hotel_Mode)
        menubar.add_cascade(label="Restaurant Manager", menu=Restaurantmenu)

        Salesmenu = Menu(menubar, tearoff=0)
        submenu = Menu(Salesmenu)
        submenu.add_command(label="Excel", command=ESC)
        submenu.add_command(label="CSV", command=donothing)
        submenu.add_command(label="Print", command=donothing)

        Salesmenu.add_command(label="Sales", command=CRJ1)
        Salesmenu.add_separator()
        Salesmenu.add_cascade(label="Export", menu=submenu)
        Salesmenu.add_separator()
        Salesmenu.add_command(label="Search Sales By Date", command=donothing)
        Salesmenu.add_command(label="Clear Sales", command=donothing)
        menubar.add_cascade(label="Sales Data", menu=Salesmenu)


        helpmenu = Menu(menubar, tearoff=0)
        helpmenu.add_command(label="Help Index", command=Help)
        helpmenu.add_command(label="About...", command=donothing)
        helpmenu.add_command(label="Yoco", command=Open_Yoco)
        helpmenu.add_command(label="Donate", command=Donate_Yoco)
        menubar.add_cascade(label="Help", menu=helpmenu)

        Page1.config(menu=menubar)

        F1 = Frame(Page1, height=100, width=(Page1.winfo_width()), bg="#E9E9E9", relief="raise")
        F1.grid(row=0,column=0)
        F1.grid_propagate(0)

        Black_F = Frame(Page1, height=23, width=(Page1.winfo_width()), bg="#AAAAAA", relief="raise")
        Black_F.grid(row=1,column=0,sticky='n')

        Label(F1, text=("Made By Connor Hess  V" + str(Version)), fg="white", bg="gray").place(x=1375,y=1)

        F1_1 = LabelFrame(F1, text="Systems", bg="#E9E9E9", relief="raise")
        F1_1.grid(row=0,column=0)

        LEVEL1 = Button(F1_1, text="Barcode Cart", width=13, height=1, fg="white", bg="green", command=Cart1, bd=2).grid(row=0,column=0)
        LEVEL2 = Button(F1_1, text="Restorant System", width=13, height=1, fg="white", bg="green", command=RSYS, bd=2).grid(row=1,column=0)


    elif LEVL == 2:
        menubar = Menu(Page1)
        filemenu = Menu(menubar, tearoff=0)
        filemenu.add_command(label="Exit", command=Page1_close)
        menubar.add_cascade(label="File", menu=filemenu)

        Usermenu = Menu(menubar, tearoff=0)
        Usermenu.add_command(label="User List", command=donothing)
        Usermenu.add_separator()
        Usermenu.add_separator()
        menubar.add_cascade(label="User's", menu=Usermenu)

        Posmenu = Menu(menubar, tearoff=0)
        Posmenu.add_command(label="Product List", command=P_List)
        Posmenu.add_separator()
        Posmenu.add_command(label="New Product", command=New_product)
        Posmenu.add_command(label="Delete Product", command=Del_product)
        Posmenu.add_separator()
        Posmenu.add_command(label="Add Scale Item", command=donothing)
        Posmenu.add_command(label="Delete Scale Item", command=donothing)
        menubar.add_cascade(label="POS System", menu=Posmenu)

        Hotel_Mode = IntVar()
        Restaurantmenu = Menu(menubar, tearoff=0)
        Restaurantmenu.add_command(label="Item List", command=donothing)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_command(label="New Item", command=Add_Menu1)
        Restaurantmenu.add_command(label="Delete Item", command=Add_Menu1)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_command(label="Add Menu category", command=donothing)
        Restaurantmenu.add_command(label="Remove Menu category", command=donothing)
        Restaurantmenu.add_separator()
        Restaurantmenu.add_checkbutton(label="Hotel Mode", onvalue=1, offvalue=0, variable=Hotel_Mode)
        menubar.add_cascade(label="Restaurant Manager", menu=Restaurantmenu)

        Salesmenu = Menu(menubar, tearoff=0)
        submenu = Menu(Salesmenu)
        submenu.add_command(label="Excel", command=ESC)
        submenu.add_command(label="CSV", command=donothing)
        submenu.add_command(label="Print", command=donothing)

        Salesmenu.add_command(label="Sales", command=CRJ1)
        Salesmenu.add_separator()
        Salesmenu.add_cascade(label="Export", menu=submenu)
        Salesmenu.add_separator()
        Salesmenu.add_command(label="Search Sales By Date", command=donothing)
        menubar.add_cascade(label="Sales Data", menu=Salesmenu)


        helpmenu = Menu(menubar, tearoff=0)
        helpmenu.add_command(label="Help Index", command=Help)
        helpmenu.add_command(label="About...", command=donothing)
        menubar.add_cascade(label="Help", menu=helpmenu)

        Page1.config(menu=menubar)

        F1 = Frame(Page1, height=100, width=(Page1.winfo_width()), bg="#E9E9E9", relief="raise")
        F1.grid(row=0,column=0)
        F1.grid_propagate(0)

        Black_F = Frame(Page1, height=4, width=(Page1.winfo_width()), bg="#AAAAAA", relief="raise")
        Black_F.grid(row=1,column=0,sticky='n')

        Label(F1, text=("Made By Connor Hess  V" + str(Version)), fg="white", bg="gray").place(x=1300,y=1)


        F1_1 = LabelFrame(F1, text="Systems", bg="#E9E9E9", relief="raise")
        F1_1.grid(row=0,column=0)

        LEVEL1 = Button(F1_1, text="Barcode Cart", width=12, height=1, fg="white", bg="green", command=Cart1, bd=2).grid(row=0,column=0)
        LEVEL2 = Button(F1_1, text="Restorant System", width=12, height=1, fg="white", bg="green", command=RSYS, bd=2).grid(row=1,column=0)


    elif LEVL == 3:
        menubar = Menu(Page1)
        filemenu = Menu(menubar, tearoff=0)
        filemenu.add_command(label="Exit", command=Page1_close)
        menubar.add_cascade(label="File", menu=filemenu)

        Usermenu = Menu(menubar, tearoff=0)
        Usermenu.add_command(label="User List", command=donothing)
        menubar.add_cascade(label="User's", menu=Usermenu)

        Posmenu = Menu(menubar, tearoff=0)
        Posmenu.add_command(label="Product List", command=P_List)
        menubar.add_cascade(label="POS System", menu=Posmenu)

        Hotel_Mode = IntVar()
        Restaurantmenu = Menu(menubar, tearoff=0)
        Restaurantmenu.add_command(label="Item List", command=donothing)
        menubar.add_cascade(label="Restaurant Manager", menu=Restaurantmenu)

        helpmenu = Menu(menubar, tearoff=0)
        helpmenu.add_command(label="Help Index", command=Help)
        helpmenu.add_command(label="About...", command=donothing)
        menubar.add_cascade(label="Help", menu=helpmenu)

        Page1.config(menu=menubar)

        F1 = Frame(Page1, height=100, width=(Page1.winfo_width()), bg="#E9E9E9", relief="raise")
        F1.grid(row=0,column=0)
        F1.grid_propagate(0)

        Black_F = Frame(Page1, height=4, width=(Page1.winfo_width()), bg="#AAAAAA", relief="raise")
        Black_F.grid(row=1,column=0,sticky='n')

        Label(F1, text=("Made By Connor Hess  V" + str(Version)), fg="white", bg="gray").place(x=1300,y=1)

        F1_1 = LabelFrame(F1, text="Systems", bg="#E9E9E9", relief="raise")
        F1_1.grid(row=0,column=0)


        LEVEL1 = Button(F1_1, text="Barcode Cart", width=12, height=1, fg="white", bg="green", command=Cart1, bd=2).grid(row=0,column=0)
        LEVEL2 = Button(F1_1, text="Restorant System", width=12, height=1, fg="white", bg="green", command=RSYS, bd=2).grid(row=1,column=0)





    m1 = PanedWindow(Page1, bg="grey", height=(Page1.winfo_height() - 150), width=(Page1.winfo_width()))
    m1.grid(row=2,column=0)

    F2 = Frame(m1, width=180, bg="#E9E9E9", relief="raise")
    F2.grid_propagate(0)
    m1.add(F2)


    m2 = PanedWindow(m1, sashwidth=8, orient=VERTICAL, bg="grey")
    m1.add(m2)

    F3 = Frame(m2, width=250, height=380 , bg="#E9E9E9", relief="raise")
    F3.grid_propagate(0)
    m2.add(F3)


    tabControl = ttk.Notebook(F3)

    tab1 = ttk.Frame(tabControl)
    tab2 = ttk.Frame(tabControl)
    tab3 = ttk.Frame(tabControl)


    tabControl.add(tab1, text='30 Day Income')
    tabControl.add(tab2, text='30 Day Meat Sales')
    tabControl.add(tab3, text='Statistics')



##    F5 = Frame(tab1, width=250, height=50 , bg="#E9E9E9", relief="raise")
##    F5.grid_propagate(0)
##    F5.grid(row=0,column=0,sticky='w')


    var = StringVar()
    label2 = Label(F1, textvariable=var, relief=RAISED )
    var.set((time.strftime("%H:%M")))

    Message_var = StringVar()
    label3 = Label(Black_F, textvariable=Message_var,width=230, anchor='w' , relief=RAISED )
    Message_var.set("...")



##    Month_query = Entry(F5, bd=2)
##    Month_query.grid(row=0,column=0)



    F6 = Frame(tab1, bg="#E9E9E9", relief="raise")
    F6.grid_propagate(0)
    F6.grid(row=0,column=0)

    F8 = Frame(tab2, width=250, height=250 , bg="#E9E9E9", relief="raise")
    F8.grid_propagate(0)
    F8.grid(row=1,column=0)

    F10 = Frame(tab3, bg="#E9E9E9", relief="raise")
##    F10.grid_propagate(0)
    F10.grid(row=1,column=0)

##    Row_Inc = 0
##    Col_Inc = 1
##    Col_Inc_2 = 0
##    Col_Inc_3 = 3
##    Col_Inc_4 = 2
##
##    c.execute("SELECT * FROM Scale")
##    for row in c.fetchall():
##        Total_Meat_Stats = 0
##        PPK = (float(row[2]))
##        if Row_Inc >= 20:
##            Col_Inc += 4
##            Row_Inc = 0
##            Col_Inc_2 += 4
##            Col_Inc_3 += 4
##            Col_Inc_4 += 4
##
##        Stats = StringVar()
##        label6 = Label(F10, textvariable=Stats, anchor='w', pady=4)
##        label6.grid(row=Row_Inc,column=Col_Inc,sticky='w')
##
##        label7 = Label(F10, text=' |  Income: R', anchor='e', pady=4)
##        label7.grid(row=Row_Inc,column=Col_Inc_4,sticky='e')
##
##        Stats_Inc = StringVar()
##        label8 = Label(F10, textvariable=Stats_Inc, anchor='w', pady=4, padx=4)
##        label8.grid(row=Row_Inc,column=Col_Inc_3,sticky='w')
##
##        label9 = Label(F10, text=(row[1]) + ' : kg ', anchor='e', pady=4)
##        label9.grid(row=Row_Inc,column=Col_Inc_2,sticky='e')
##        Row_Inc += 1
##
##        c.execute("SELECT * FROM Sales WHERE Item=?",((row[1]),))
##        for row in c.fetchall():
##            Total_Meat_Stats += float(row[9])
##
##        Stats.set(Total_Meat_Stats)
##        Stats_Inc.set(round((float(Total_Meat_Stats) * PPK),2))


    def Stats_generate():
        label6 = Label(F10, text='', anchor='w', pady=4)
        label6.grid(row=0,column=0,sticky='w')

        label7 = Label(F10, text='', anchor='e', pady=4)
        label7.grid(row=0,column=0,sticky='e')

        label8 = Label(F10, text='', anchor='w', pady=4, padx=4)
        label8.grid(row=0,column=0,sticky='w')

        label9 = Label(F10, text='', anchor='e', pady=4)
        label9.grid(row=0,column=0,sticky='e')

        label6.destroy()
        label7.destroy()
        label8.destroy()
        label9.destroy()
        Row_Inc = 0
        Col_Inc = 1
        Col_Inc_2 = 0
        Col_Inc_3 = 3
        Col_Inc_4 = 2

        c.execute("SELECT * FROM Scale")
        for row in c.fetchall():
            Total_Meat_Stats = 0
            PPK = (float(row[2]))
            if Row_Inc >= 20:
                Col_Inc += 4
                Row_Inc = 0
                Col_Inc_2 += 4
                Col_Inc_3 += 4
                Col_Inc_4 += 4

            Stats = StringVar()
            label6 = Label(F10, textvariable=Stats, anchor='w', pady=4)
            label6.grid(row=Row_Inc,column=Col_Inc,sticky='w')

            label7 = Label(F10, text=' |  Income: R', anchor='e', pady=4)
            label7.grid(row=Row_Inc,column=Col_Inc_4,sticky='e')

            Stats_Inc = StringVar()
            label8 = Label(F10, textvariable=Stats_Inc, anchor='w', pady=4, padx=4)
            label8.grid(row=Row_Inc,column=Col_Inc_3,sticky='w')

            label9 = Label(F10, text=(row[1]) + ' : kg ', anchor='e', pady=4)
            label9.grid(row=Row_Inc,column=Col_Inc_2,sticky='e')
            Row_Inc += 1

            c.execute("SELECT * FROM Sales WHERE Item=?",((row[1]),))
            for row in c.fetchall():
                Total_Meat_Stats += float(row[9])

            Stats.set(Total_Meat_Stats)
            Stats_Inc.set(round((float(Total_Meat_Stats) * PPK),2))

    Stats_generate()


    fig = Figure(figsize=(15, 7), dpi=85)
    fig.subplots_adjust(left=0.05, bottom=0.08, right=0.98, top=0.95, wspace=None, hspace=None)
    Fig_plot_1 = fig.add_subplot(111)


    fig_2 = Figure(figsize=(15, 7), dpi=85)
    fig_2.subplots_adjust(left=0.05, bottom=0.08, right=0.98, top=0.95, wspace=None, hspace=None)
    Fig_plot_2 = fig_2.add_subplot(111)

    def daterange(start_date, end_date):
        for n in range(int ((end_date - start_date).days)):
            yield start_date + timedelta(n)

    def animate(i=1):
        var.set((time.strftime("%H:%M")))
        Fig_plot_1.clear()
        fig.suptitle('Income comparison')
        Fig_plot_1.set_ylabel('Income')
        Fig_plot_1.set_xlabel('Date')
        xList = []
        yList = []

        end_date = datetime.today() + timedelta(days=1)
        start_date = datetime.today() - timedelta(days=30)
        for single_date in daterange(start_date, end_date):
##            print(single_date.strftime("%d-%m-%Y"))
            Date_query = (single_date.strftime("%d-%m-%Y"))
            Date_query_show = (single_date.strftime("%d-%m"))
            Total_Day = 0
            c.execute('SELECT Date, Bank FROM CRJ WHERE Date=?',(Date_query,))
            for row in c.fetchall():
##                print(row)
                Total_Day += row[1]
            xList.append(Date_query_show)
            yList.append(Total_Day)
        Fig_plot_1.plot(xList,yList)
        try:
            response = requests.get('https://raw.githubusercontent.com/connorhess/EasyPOS/master/Message.txt')
            Message = response.text
            Message_var.set(Message)
        except:
            Message_var.set("Ofline")

        Stats_generate()

    def animate2(i=1):
        Fig_plot_2.clear()
        fig_2.suptitle('Meat In KG')
        Fig_plot_2.set_ylabel('Meat Sold (KG)')
        Fig_plot_2.set_xlabel('Date')
        xList2 = []
        yList2 = []
        end_date = datetime.today() + timedelta(days=1)
        start_date = datetime.today() - timedelta(days=30)
        for single_date in daterange(start_date, end_date):
            Date_query = (single_date.strftime("%d-%m-%Y"))
            Date_query_show = (single_date.strftime("%d-%m"))
            Total_Day_W = 0
            c.execute('SELECT Date, Weight FROM Sales WHERE Date=?',(Date_query,))
            for row in c.fetchall():
##                print(row)
                Total_Day_W += row[1]
            xList2.append(Date_query_show)
            yList2.append(Total_Day_W)
        Fig_plot_2.plot(xList2,yList2)




##    Button(F5, text="Refresh", width=12, height=1, fg="white", bg="green", command=animate, bd=2).grid(row=0,column=1)




    canvas = FigureCanvasTkAgg(fig, master=F6)  # A tk.DrawingArea.
    canvas.draw()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
    toolbar = NavigationToolbar2Tk(canvas, F6)
##    toolbar.update()
    canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)



    canvas2 = FigureCanvasTkAgg(fig_2, master=F8)  # A tk.DrawingArea.
    canvas2.draw()
    canvas2.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
    toolbar = NavigationToolbar2Tk(canvas2, F8)
##    toolbar.update()
    canvas2.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)

    label2.place(x=1500,y=30)
    label3.place(x=0,y=0)



    def Re_Fresh():
        animate()
        animate2()
        Stats_generate()

    tabControl.pack(expand=1, fill="both")

    LEVEL3 = Button(F2, text="Refresh", width=12, height=1, fg="white", bg="green", command=Re_Fresh, bd=2).grid(row=0,column=0)


    ani = animation.FuncAnimation(fig, animate, interval=60000)
    ani2 = animation.FuncAnimation(fig_2, animate2, interval=60000)

##    Page1.columnconfigure(0, weight=1)
##    Page1.rowconfigure(0, weight=0) # not needed, this is the default behavior
##    Page1.rowconfigure(1, weight=1)
##    Page1.rowconfigure(2, weight=1)
    Page1.mainloop()


##RUN1("Connor")

#state=DISABLED
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================
#========================================================================================================================================================================================================================================


