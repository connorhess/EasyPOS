import sqlite3
from tkinter import *
from tkinter import filedialog
import time
import random
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl import load_workbook
import tempfile
import os
from functools import partial
from prettytable import PrettyTable
from tkinter.font import Font
from os.path import join as pjoin
import Error
from PIL import ImageTk,Image


button_1 = "Beers & Ciders"
button_2 = "wine/per Glass"
button_3 = "Soft Drinks"
button_4 = "Gin"
button_5 = "Spirits"
button_6 = "Whiskey"
button_7 = "Brandy"



try:
    filepath="CRJ.xlsx"
    # load demo.xlsx 
    wb=load_workbook(filepath)
except:
    # set file path
    filepath="CRJ.xlsx"
    # load demo.xlsx
    wb=Workbook()
    # save workbook 
    wb.save(filepath)

    
TIME = (time.strftime("%H:%M"))
DATE = (time.strftime("%d-%m-%Y"))
Day = (time.strftime("%d"))
Month = (time.strftime("%m"))
Year = (time.strftime("%Y"))

conn = sqlite3.connect('Shop_Database.db')
c = conn.cursor()


x = PrettyTable()

##conn = sqlite3.connect('Restorant.db')
##c = conn.cursor()


#RSYSS.TableNumberGenerate("Table TK var","title", )


##def GENTBS():
def PASS(Table,NO):
    TableNumberGenerate(Table,NO)

def Config1():
    c.execute('''CREATE TABLE IF NOT EXISTS Tables(Item_PLU REAL, Item_Name TEXT, Item_Price INT, QTY INT, FPrice INT, Table_No INT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS Starters(Item_PLU REAL, Item_Name TEXT, Item_Price REAL, Menu INT)''')

def TableNumberAdd(B1,B2,B3,B4,TableNO,B6):
##    B1 = 
##    B2 = 
##    B3 = 
##    B4 = 
    B5 = TableNO
    
    c.execute('''INSERT INTO Tables(Item_PLU, Item_Name, Item_Price, QTY, Table_No, FPrice) VALUES(?, ? ,? ,? ,? ,?)''',(B1, B2, B3, B4, B5, B6))
    conn.commit()


def ItemAdd(B1,B2,B3,B4):
##    B1 = 
##    B2 = 
##    B3 = 
##    B4 = 
    c.execute('''INSERT INTO Starters(Item_PLU, Item_Name, Item_Price, Menu) VALUES(?, ? ,? ,?)''',(B1, B2, B3, B4))
    conn.commit()

    

def PASS():
    pass





    

def print_all():
##    Path_1 = filedialog.asksaveasfile(initialdir = "/",title = "select file",filetypes = (("all files",".")))
##    print(Path_1)

    Path_1 =  filedialog.askdirectory(initialdir = "/",title = "Select file")
    print (Path_1)

    
    
    x.field_names = ["PLU","Item", "Price", "QTY", "Total Price", "Room Number",]
    x.clear_rows()
    
    
    total_add_up = 0
    cashup_total = 0
    
    Rooms = [4,5,6,7,8,9,10,11,12,14,15,16,17,18,19,20,21,22,201,202,203,204,1001,1002,1003]
    
    for table_no in Rooms:
        F_Name = "Room-" + str(table_no) + "_Print_" + str(time.strftime("%d-%m-%Y"))
        Path_2 = (str(Path_1) + (F_Name)) + ".txt"
        print(Path_2)
        F = open(Path_2, "a")
        total_add_up = 0
        x.clear_rows()
        c.execute("SELECT Item_PLU,Item_Name,Item_Price,QTY,FPrice,Table_No FROM Tables WHERE Table_No=?",(table_no,))
        TBL = c.fetchall()
        for row in TBL:
            total_add_up = total_add_up + (row[4])
            E = (row)
            x.add_row(E)
##        print(x)
        F.write((str(x)) + "\ntotal: " + str(total_add_up) + "\n=========================================================\n\n")
        F.close()


        
    F = open(Path_2, "r")
    print(F.read())

    filename = tempfile.mktemp(".txt")
    
    open (filename, "w").write(F.read())
##    os.startfile(filename, "print")

    print(F.read())
    F.close() 

#56

##print_all()

##
##TableNumberGenerate('Table 1', '1')
PE = "Pastaaaa"
PEE = "2"
##TableNumberAdd('2','pizzaaa','50.00','1','2')
##ItemAddS(PEE,PE,20.00)
##ItemAddM(PEE,PE,20.00)
##ItemAddD(PEE,PE,20.00)
##ItemAddW(PEE,PE,20.00)
##ItemAddB(PEE,PE,20.00)
##ItemAddBE(PEE,PE,20.00)

##.transient([PageR])

def RSYS():
    PageR = Tk()
    PageR.title("Item Database")
    PageR.configure(background="orange")
    ##        PageR.geometry("1100x800")
    PageR.attributes("-fullscreen", True)

    def exit_PageR():
        PageR.destroy()

    def TableNumberGenerate(Title,tableNO):
        conn = sqlite3.connect('Shop_Database.db')
        c = conn.cursor()
        
        c.execute('''CREATE TABLE IF NOT EXISTS Tables(Item_PLU REAL, Item_Name TEXT, Item_Price INT, QTY INT, FPrice INT, Table_No INT)''')
        c.execute('''CREATE TABLE IF NOT EXISTS Starters(Item_PLU REAL, Item_Name TEXT, Item_Price REAL, Menu INT)''')
        a1 = Toplevel()
        a1.title(Title)
        a1.configure(background="cyan")
        a1.geometry("500x300")
        a1.transient([PageR])

        top = a1
        
        Lb1 = Listbox(top, width=30)
        Lb1.insert(1, "Item                   Price           QTY")


        def BPAGE(Name,Menu):
            a2 = Toplevel()
            a2.title(Name)
            a2.configure(background="cyan")
            a2.geometry("850x600")
            a2.transient([PageR])
            L1 = Label(a2, text="QTY    [^]Up= 2    [v]Down= 1",font=Font(family='Helvetica',size=40,weight='bold'))
            L1.grid(row=1,column=2)
            w1 = Spinbox(a2, from_=1, to=20, font=Font(family='Helvetica',size=40,weight='bold'), width=4)
            w1.grid(row=1,column=1)
            QQTY = w1.get()
            def CLOSE_BPAGE():
                a2.destroy()
            b1 = Button(a2, text ='Close', command=CLOSE_BPAGE)
            b1.grid(row=1,column=3)
            def add1(IName,QQTY):
                v2 = IName
                print (v2)
                c.execute("SELECT * FROM Starters WHERE Item_Name=? ",(v2,))
                DA22 = c.fetchall()
                for row in DA22:
                    QQTY = w1.get()
                    Lb1.insert(1, (str(row[1]) + ("  ") + (str(row[2])) + ("  ") + str(QQTY)))
                    print (row)
                    B1 = (row[0])
                    B2 = (row[1])
                    B3 = (row[2])
                    B4 = (w1.get())
                    B6 = (row[2]) * int(w1.get())
                    TableNumberAdd(B1,B2,B3,B4,tableNO,B6)
                    a2.destroy()

            c.execute("SELECT * FROM Starters WHERE Menu=?",(Menu,))
            DA23 = c.fetchall()
            YU = 30
            XU = 0
            POSS = 1
            W = 20
            H = 3
            for row in DA23:
                if POSS == 1:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=2,column=1)
                    POSS = POSS + 1
                elif POSS == 2:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=3,column=1)
                    POSS = POSS + 1
                elif POSS == 3:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=4,column=1)
                    POSS = POSS + 1
                elif POSS == 4:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=5,column=1)
                    POSS = POSS + 1
                elif POSS == 5:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=6,column=1)
                    POSS = POSS + 1
                elif POSS == 6:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=7,column=1)
                    POSS = POSS + 1
                elif POSS == 7:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=8,column=1)
                    POSS = POSS + 1
                    
                elif POSS == 8:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=2,column=2)
                    POSS = POSS + 1
                elif POSS == 9:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=3,column=2)
                    POSS = POSS + 1
                elif POSS == 10:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=4,column=2)
                    POSS = POSS + 1
                elif POSS == 11:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=5,column=2)
                    POSS = POSS + 1
                elif POSS == 12:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=6,column=2)
                    POSS = POSS + 1
                elif POSS == 13:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=7,column=2)
                    POSS = POSS + 1
                elif POSS == 14:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=8,column=2)
                    POSS = POSS + 1

                elif POSS == 15:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=2,column=3)
                    POSS = POSS + 1
                elif POSS == 16:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=3,column=3)
                    POSS = POSS + 1
                elif POSS == 17:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=4,column=3)
                    POSS = POSS + 1
                elif POSS == 18:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=5,column=3)
                    POSS = POSS + 1
                elif POSS == 19:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=6,column=3)
                    POSS = POSS
                elif POSS == 20:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=7,column=3)
                    POSS = POSS + 1
                elif POSS == 21:
                    print (row[1])
                    b3 = Button(a2, text =(row[1]), width=W, height=4, command=partial(add1,(row[1]), w1.get()))
                    b3.grid(row=8,column=3)
                    POSS = POSS + 1

        
        c.execute("SELECT * FROM Tables WHERE Table_No=?",(tableNO,))
        DA22 = c.fetchall()
        for row in DA22:
            S1 = ((25 - (len(str(row[1])))) * " ")
            S2 = ((15 - (len(str(row[2])))) * " ")
            Lb1.insert(1, ((str(row[1])) + S1 + (str(row[2])) + S2 + (str(row[3]))))


        def F9():
            print(Lb1.get(0, last = 100))

            Page4 = Toplevel()
            Page4.title("Shop Database")
            Page4.geometry("400x300")
            Page4.transient([PageR])
            Label(Page4, text="Total", bd=2).place(x=1,y=1)
            c.execute("SELECT SUM(FPrice) FROM Tables WHERE Table_No=?",(tableNO,))
            DA23 = c.fetchall()
            for row in DA23:
                global SUM
                SUM = (row[0])            
            Label(Page4, text=(SUM), bd=2).place(x=60,y=1)
            Label(Page4, text="Cash", bd=2).place(x=1,y=30)
            text2 = Text(Page4, height=1, width=10)

            #======Numpad======#
            def K9():
                text2.insert(INSERT, "9")
            def K8():
                text2.insert(INSERT, "8")
            def K7():
                text2.insert(INSERT, "7")
            def K6():
                text2.insert(INSERT, "6")
            def K5():
                text2.insert(INSERT, "5")
            def K4():
                text2.insert(INSERT, "4")
            def K3():
                text2.insert(INSERT, "3")
            def K2():
                text2.insert(INSERT, "2")
            def K1():
                text2.insert(INSERT, "1")
            def K0():
                text2.insert(INSERT, "0")
            def K():
                text2.insert(INSERT, ".")
            def K10():
                text2.delete(1.0, 1000.0)
                
            F102 = Frame(Page4, bd=8, bg="grey", relief="raise")
            F102.place(x=150,y=1)

            scale_1 = 3

            W = 2*scale_1
            H = 1*scale_1

            Button(F102, text="9", width=W, height=H, bg="blue", fg="white", command=K9, bd=2).grid(row=1,column=2)
            Button(F102, text="8", width=W, height=H, bg="blue", fg="white", command=K8, bd=2).grid(row=1,column=1)
            Button(F102, text="7", width=W, height=H, bg="blue", fg="white", command=K7, bd=2).grid(row=1,column=0)
            Button(F102, text="6", width=W, height=H, bg="blue", fg="white", command=K6, bd=2).grid(row=2,column=2)
            Button(F102, text="5", width=W, height=H, bg="blue", fg="white", command=K5, bd=2).grid(row=2,column=1)
            Button(F102, text="4", width=W, height=H, bg="blue", fg="white", command=K4, bd=2).grid(row=2,column=0)
            Button(F102, text="3", width=W, height=H, bg="blue", fg="white", command=K3, bd=2).grid(row=3,column=2)
            Button(F102, text="2", width=W, height=H, bg="blue", fg="white", command=K2, bd=2).grid(row=3,column=1)
            Button(F102, text="1", width=W, height=H, bg="blue", fg="white", command=K1, bd=2).grid(row=3,column=0)
            Button(F102, text="0", width=W, height=H, bg="blue", fg="white", command=K0, bd=2).grid(row=4,column=1)
            Button(F102, text=".", width=W, height=H, bg="blue", fg="white", command=K, bd=2).grid(row=4,column=0)
            Button(F102, text="C", width=W, height=H, bg="blue", fg="white", command=K10, bd=2).grid(row=4,column=2)

            text2.place(x=60,y=30)


            
            def F10():
                global SUM
                EE2 = (text2.get(1.0, 1000.0))
                CHANGE = float(EE2) - float(SUM)
                if float(EE2) < float(SUM):
                    print("error 1")
                    MsgBox_001 = messagebox.showerror ('ERROR',Error.Error_001,icon = Error.Error_icon)

                else:
                    Page4.destroy()
                    def OK1():
                        c.execute("DELETE FROM Tables WHERE Table_No=?",(tableNO,))
                        conn.commit()
                        Lb1.delete(0, 200)
                        a1.destroy()
                        Page5.destroy()    
                        Button(Page5, text="Print", width=15, height=1, fg="white", bg="green", command=iprint, bd=2).place(x=1,y=140)
                        c.execute("SELECT SUM(FPrice) FROM Tables WHERE Table_No=?",(tableNO,))
                        DA23 = c.fetchall()
                        for row in DA23:
                            SUM = (row[0])
                            print (row)
                            RR1 = random.randint(1,1000000000)
                            Bb1 = RR1
                            Bb2 = TimeDate
                            Bb3 = "Table" + str(tableNO)
                            Bb4 = str(EE2)
                            Bb5 = str(SUM)
                            Bb6 = str(Lb1.get(1, 200))
                            Bb7 = 1
                            c.execute('''INSERT INTO CRJ(ID, Day, Month, Year, Time, Description, Amount, Bank, Item, QTY) VALUES(?, ?, ?, ?, ? ,? ,? ,? ,? ,?)''',(Bb1, (time.strftime("%d")), (time.strftime("%m")),(time.strftime("%Y")),(time.strftime("%H:%M")), Bb3, Bb4, Bb5, Bb6, Bb7))
                            conn.commit()
                    
                    SUM = str(SUM)
                    Page5 = Toplevel()
                    Page5.title("Shop Database")
                    Page5.configure(background="grey")
                    Page5.geometry("400x300")
                    Page5.transient([PageR])
                    
                    
                    Label(Page5, text=(CHANGE), bd=2).place(x=100,y=1)
                    def iprint():
                        M = Lb1.get(0, last = 1000)
                        print(Lb1.get(0, last = 1000))

                        
                        Q = ("\n\nThank you for staying by us.                  Software Made By Connor Hess\n       Mimosa Lodge\nRoom Number: " + tableNO + "\n")
                        filename = tempfile.mktemp(".txt")
                        open (filename, "w").write(Q)
                        
                        
                        c.execute("SELECT Item_PLU,Item_Name,Item_Price,QTY FROM Tables WHERE Table_No=?",(tableNO,))
                        TBL = c.fetchall()
                        x.clear_rows()
                        for row in TBL:
                            x.field_names = ["PLU","Item", "Price", "QTY"]
                            E = (row)
                            x.add_row(E)
                        print(x)
                        open (filename, "a").write(str(x))

                        open (filename, "a").write("\nChange: " + str(CHANGE) + "        Total: " + str(SUM) + "    Ammount recieved: " + str(EE2))
                        def RPrint():
                            os.startfile(filename, "print")

                        Button(Page5, text="RePrint", width=15, height=1, fg="white", bg="green", command=RPrint, bd=2).place(x=1,y=165)

                        os.startfile(filename, "print")
                    
                    
                    
                    
                        
                    Button(Page5, text="OK", width=20, height=2, fg="white", bg="green", command=OK1, bd=2).place(x=1,y=70)
                
                

            Button(Page4, text="PAY", width=20, height=1, fg="white", bg="green", command=F10, bd=2).place(x=1,y=60)

        def Other_price():
            a5 = Toplevel()
            a5.title(Title)
            a5.configure(background="cyan")
            a5.geometry("500x300")
            a5.transient([PageR])

            
            
        def F1():
            v9, v10, v11 = (Lb1.get(Lb1.curselection())).split()
            print (v9)
            print(Lb1.curselection())
            B1 = v9
            c.execute("DELETE FROM Tables WHERE Item_Name=? AND QTY=?",(B1,v11))
            conn.commit()
            Lb1.delete(Lb1.curselection())
        def F2():
            MsgBox = messagebox.askquestion ('Warning','Are you sure you want to clear\nThis will delete everything and can not be undone',icon = 'warning')
            if MsgBox == 'yes':
                Lb1.delete(1, 200)
                c.execute("DELETE FROM Tables WHERE Table_No=?",(tableNO,))
                conn.commit()

        def Close_R():
            top.destroy()
        
        b1 = Button(top, text ="Delete selected", command=F1)
        b1.place(x=1,y=1)

        b2 = Button(top, text ="Delete All", command=F2)
        b2.place(x=90,y=1)

        b4 = Button(top, text ="Pay", command=F9)
        b4.place(x=200,y=1)

        b4 = Button(top, text ="Close", command=Close_R)
        b4.place(x=1,y=250)
        
    ##    c.execute("SELECT * FROM Settings WHERE ID=7")
    ##    for row in c.fetchall():
    ##        print(row[2])
    ####        ADA = (row[2])
    ##        ADA = "2"
    ##        if ADA == "2":


    ##        def F3():
    ##            a2 = Toplevel()
    ##            a2.title("Starters")
    ##            a2.configure(background="cyan")
    ##            a2.geometry("400x300")
    ##            L1 = Label(a2, text="QTY")
    ##            L1.place(x=50,y=1)
    ##            w1 = Spinbox(a2, from_=1, to=20, width=4)
    ##            w1.place(x=80,y=1)
    ##            def add1():
    ##                Lb1.insert(1, (Lb2.get(Lb2.curselection())) + (str(w1.get())))
    ##                look1 = (Lb2.get(Lb2.curselection()) + (str(w1.get())))#[:5]
    ##                v2, v3, v4 = look1.split()
    ##                print (v2)
    ##                c.execute("SELECT * FROM Starters WHERE Item_Name=? ",(v2,))
    ##                DA22 = c.fetchall()
    ##                for row in DA22:
    ##                    print (row)
    ##                    B1 = (row[0])
    ##                    B2 = (row[1])
    ##                    B3 = (row[2])
    ##                    B4 = (w1.get())
    ##                    B6 = (row[2]) * int(w1.get())
    ##                    TableNumberAdd(B1,B2,B3,B4,tableNO,B6)
    ##                    a2.destroy()
    ##            Lb2 = Listbox(a2, width=40)
    ##            c.execute("SELECT * FROM Starters WHERE Menu=1")
    ##            DA23 = c.fetchall()
    ##            for row in DA23:
    ##                S3 = ((14 - (len(str(row[1])))) * " ")
    ##                S4 = ((10 - (len(str(row[2])))) * " ")
    ##                Lb2.insert(1, ((str(row[1])) + S3 + "  R"  + (str(row[2])) + S4))
    ##                
    ##            Lb2.place(x=1,y=30)
    ##            b3 = Button(a2, text ="Add", command=add1)
    ##            b3.place(x=1,y=1)
                
        def F3():
            BPAGE(button_1,"1")
            
        def F4():
            BPAGE(button_2,"2")
            
        def F5():
            BPAGE(button_3,"3")
        
        def F6():
            BPAGE(button_4,"4")
        
        def F7():
            BPAGE(button_5,"5")

        def F8():
            BPAGE(button_6,"6")

        def F31():
            BPAGE(button_7,"7")

        

        X1 = 250
        X2 = 50
        X3 = 250 + 150

        button_number = 1


        c.execute("SELECT * FROM MenuS ORDER BY Menu_Number ASC")
        for row in c.fetchall():
            if button_number == 1:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 2:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 3:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 4:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 5:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 6:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 7:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X1,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
                
            elif button_number == 8:
                X2 = 50
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 9:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 10:
                bu1 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                bu1.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 11:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 12:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 13:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
            elif button_number == 14:
                b3 = Button(top, text =(row[0]), command=partial(BPAGE,(row[0]),(row[1])))
                b3.place(x=X3,y=X2)
                button_number = button_number + 1
                X2 = X2 + 30
                
            
    ##    b3 = Button(top, text =button_1, command=F3)
    ##    b3.place(x=X1,y=50)

    ##    b4 = Button(top, text =button_2, command=F4)
    ##    b4.place(x=X1,y=80)
    ##
    ##    b5 = Button(top, text =button_3, command=F5)
    ##    b5.place(x=X1,y=110)
    ##
    ##    b6 = Button(top, text =button_4, command=F6)
    ##    b6.place(x=X1,y=140)
    ##
    ##    b7 = Button(top, text =button_5, command=F7)
    ##    b7.place(x=X1,y=170)
    ##
    ##    b7 = Button(top, text =button_6, command=F8)
    ##    b7.place(x=X1,y=200)
    ##
    ##    b8 = Button(top, text =button_7, command=F31)
    ##    b8.place(x=X1,y=230)

    ##    b9 = Button(top, text ="subtract", command=F31)
    ##    b9.place(x=X1,y=260)
    ##
    ##    b10 = Button(top, text ="Extras", command=F31)
    ##    b10.place(x=X1,y=290)

        
        
        





        Lb1.place(x=1,y=30)


    def ROWN(PageR, ROWNUM, TBN, TBNN, LN1,LN2,LN3,LN4,LN5,LN6,LN7,LN8):
        LNV1 = TBN + (LN1)
        LNV2 = TBN + (LN2)
        LNV3 = TBN + (LN3)
        LNV4 = TBN + (LN4)
        LNV5 = TBN + (LN5)
        LNV6 = TBN + (LN6)
        LNV7 = TBN + (LN7)
        LNV8 = TBN + (LN8)

        LNVV1 = TBNN + (LN1)
        LNVV2 = TBNN + (LN2)
        LNVV3 = TBNN + (LN3)
        LNVV4 = TBNN + (LN4)
        LNVV5 = TBNN + (LN5)
        LNVV6 = TBNN + (LN6)
        LNVV7 = TBNN + (LN7)
        LNVV8 = TBNN + (LN8)
        
        RF1 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF1.grid(row=ROWNUM,column=1)
        Button(PageR, text=LNVV1, width=10, height=4, fg="black", bg="green", command=partial(TableNumberGenerate,LNV1,LN1), bd=2).grid(row=ROWNUM,column=1)

        RF2 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF2.grid(row=ROWNUM,column=2)
        Button(PageR, text=LNVV2, width=10, height=4, fg="black", bg="light green", command=partial(TableNumberGenerate,LNV2,LN2), bd=2).grid(row=ROWNUM,column=2)

        RF3 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF3.grid(row=ROWNUM,column=3)
        Button(PageR, text=LNVV3, width=10, height=4, fg="black", bg="green", command=partial(TableNumberGenerate,LNV3,LN3), bd=2).grid(row=ROWNUM,column=3)

        RF4 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF4.grid(row=ROWNUM,column=4)
        Button(PageR, text=LNVV4, width=10, height=4, fg="black", bg="light green", command=partial(TableNumberGenerate,LNV4,LN4), bd=2).grid(row=ROWNUM,column=4)

        RF5 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF5.grid(row=ROWNUM,column=5)
        Button(PageR, text=LNVV5, width=10, height=4, fg="black", bg="green", command=partial(TableNumberGenerate,LNV5,LN5), bd=2).grid(row=ROWNUM,column=5)

        RF6 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF6.grid(row=ROWNUM,column=6)
        Button(PageR, text=LNVV6, width=10, height=4, fg="black", bg="light green", command=partial(TableNumberGenerate,LNV6,LN6), bd=2).grid(row=ROWNUM,column=6)

        RF7 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF7.grid(row=ROWNUM,column=7)
        Button(PageR, text=LNVV7, width=10, height=4, fg="black", bg="green", command=partial(TableNumberGenerate,LNV7,LN7), bd=2).grid(row=ROWNUM,column=7)

        RF8 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
        RF8.grid(row=ROWNUM,column=8)
        Button(PageR, text=LNVV8, width=10, height=4, fg="black", bg="light green", command=partial(TableNumberGenerate,LNV8,LN8), bd=2).grid(row=ROWNUM,column=8)
        
    #    RF9 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
    #    RF9.grid(row=ROWNUM,column=9)
    #    Button(PageR, text=LNVV9, width=10, height=4, fg="black", bg="green", command=partial(TableNumberGenerate,LNV9,LN9), bd=2).grid(row=ROWNUM,column=9)

    #    RF10 = Frame(PageR, width=100, height=100, bd=8, bg="light grey", relief="raise")
    #    RF10.grid(row=ROWNUM,column=10)
    #    Button(PageR, text=LNVV10, width=10, height=4, fg="black", bg="light green", command=partial(TableNumberGenerate,LNV10,LN10), bd=2).grid(row=ROWNUM,column=10)


        
    Frame1 = Frame(PageR)
    Frame1.grid(row=1,column=1)

    ROWN(Frame1,"1","Room ", "Room\n","4","5","6","7","8","9","10","11")

    ROWN(Frame1,"2","Room ", "Room\n","12","14","15","16","17","18","19","20")

    ROWN(Frame1,"3","Room ", "Room\n","21","22","201","202","203","204","0","0")

    ROWN(Frame1,"4","Staff ", "Room\n","1001","1002","1003","0","0","0","0","0")

##    ROWN(PageR,"5","Table ", "Table\n","41","42","43","44","45","46","47","48","49","50")
##
##    ROWN(PageR,"6","Table ", "Table\n","41","42","43","44","45","46","47","48","49","50")
##
##    ROWN(PageR,"7","Table ", "Table\n","51","52","53","54","55","56","57","58","59","60")
##
##    ROWN(PageR,"8","Table ", "Table\n","61","62","63","64","65","66","67","68","69","70")

    Button(Frame1, text='Exit', width=10, height=4, fg="black", bg="green", command=exit_PageR, bd=2).grid(row=1,column=11)
    Button(Frame1, text='Print All', width=10, height=4, fg="black", bg="green", command=print_all, bd=2).grid(row=2,column=11)

    Frame2 = Frame(PageR)
    Frame2.grid(row=2,column=1)




##RSYS()
